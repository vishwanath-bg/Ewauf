__author__ = 'Roshanzameer M'


import json
import os
import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string
import xml.etree.ElementTree as ET
import xml.dom.minidom as minidom
from Library.Utils.Logger import Logger


class Excel_Parser:
    """
        This class helps in reading the input Excel files and return the output
        class Excel_Parser:
            def read_file
            def get_row_value
            def get_column_value
            def get_cell_value
            def write_rows
            def write_columns
            def write_cell
            def write_file
            def update_excel
            def create_excel_file
            def create_sheet
            def delete_cell
            def delete_row
            def delete_column
            def delete_sheet
            def excel_to_xml
            def excel_to_json
    """

    def __init__(self):
        self.active_sheet = None
        self.work_book = None
        self.logger = Logger()

    def read_file(self, excel_file, sheet=None):
        """
            This method reads the Excel file and also return the data
            :parameter: excel_file: specify the Excel file path
            :param sheet: specify the sheet name else active sheet will be picked
            :return: returns file data if executed else Exception
        """
        try:
            # Read the Excel file into a DataFrame
            df = pd.read_excel(excel_file, sheet_name=sheet) if sheet else pd.read_excel(excel_file)
            df = df.astype(str)
            self.logger.log_info(f"\n{df}\n")
            return df
        except Exception as error:
            self.logger.log_error(f"Error occurred while read the file: {error}\n")
            return False

    def return_values(self, data):
        """
            Extension method helps in getting column & row values from given row/column data, 
            where it checks whether list is returning None elements or cell values
            :param data: Iterator/sequence/object of retrived data
            :return: returns list of values if executed else None message
        """
        try:
            return_values = []
            if all(item.value is None for item in data):
                return 'None: None values found or No such row/column'
            for cell in data:
                if cell is not None:
                    return_values.append(cell.value)
            return return_values
        except Exception as error:
            self.logger.log_error(f"Error while extracting values: {error}\n")
            return False

    def get_row_values(self, excel_file, row_number, sheet=None):
        """
            This method fetches the single row values form the Excel sheet
            :param excel_file: Excel File path
            :param row_number: row number str/int(1, 2, 3...)
            :param sheet: specify the sheet name else active sheet will be picked
            :return: returns list of row values if executed else None message
        """
        try:
            self.work_book = openpyxl.load_workbook(excel_file)
            self.active_sheet = self.work_book[sheet] if sheet else self.work_book.active
            row = self.active_sheet[row_number]  # Retrieving the specified row
            row_values = self.return_values(row)
            return row_values
        except Exception as error:
            self.logger.log_error(f"Error occurred while fetching the row value: {error}\n")
            return False

    def get_column_values(self, excel_file, column_name, sheet=None):
        """
            This method fetches the single column values form the Excel sheet
            :param excel_file: Excel File path
            :param column_name: column name (A, B, C...)
            :param sheet: specify the sheet name else active sheet will be picked
            :return: returns list of column values if executed else None message
        """
        try:
            self.work_book = openpyxl.load_workbook(excel_file)
            self.active_sheet = self.work_book[sheet] if sheet else self.work_book.active
            column = self.active_sheet[column_name]  # Retrieving the specified column
            column_values = self.return_values(column)
            return column_values
        except Exception as error:
            self.logger.log_error(f"Error occurred while fetching the row value: {error}\n")
            return False

    def get_cell_value(self, excel_file, column_name, row_number, sheet=None):
        """
            This method fetches the cell value of mentioned row number & column name
            :param excel_file: Excel File path
            :param column_name: name of the column (A, B, C...)
            :param row_number: row number str/int(1, 2, 3...)
            :param sheet: specify the sheet name else active sheet will be picked
            :return: returns the cell value if executed else None
        """
        try:
            self.work_book = openpyxl.load_workbook(excel_file)
            self.active_sheet = self.work_book[sheet] if sheet else self.work_book.active
            column_index = column_index_from_string(column_name)
            return self.active_sheet.cell(int(row_number), column_index).value
        except Exception as error:
            self.logger.log_error(f"Error occurred while fetching the row value: {error}\n")
            return False

    def write_rows(self, excel_file, sheet, rows_data):
        """
            This method helps in writing the cell value of mentioned row number & column name
            :param excel_file: Excel file path
            :param sheet: specify the sheet name
            :param row_data: data to be written into the rows(nested list)
        """
        try:
            self.work_book = openpyxl.load_workbook(excel_file)
            self.active_sheet = self.work_book[sheet]
            for row in rows_data:
                self.active_sheet.append(row)
            self.work_book.save(excel_file)
            self.logger.log_info("Rows written successfully.\n")
            return True
        except Exception as error:
            self.logger.log_error(f"Error: {error}\n")
            return False
    
    def write_columns(self, excel_file, sheet, columns_data):
        """
            This method helps in writing the cell value of mentioned row number & column name
            :param excel_file: Excel file path
            :param sheet: specify the sheet name
            :param column_data: data to be written into the cloumns(nested list)
        """
        try:
            self.work_book = openpyxl.load_workbook(excel_file)
            self.active_sheet = self.work_book[sheet]

            rows_to_write = zip(*columns_data)
            for row in rows_to_write:
                self.active_sheet.append(row)
            self.work_book.save(excel_file)
            self.logger.log_info("Columns written successfully.\n")
            return True
        except Exception as error:
            self.logger.log_error(f"Error: {error}\n")
            return False

    def write_cell(self, excel_file, sheet, column_name, row_number, data):
        """
            This method helps in writing the cell value of mentioned row number & column name
            :param excel_file: Excel file path
            :param sheet: specify the sheet name
            :param column_name: column name (A, B, C...)
            :param row_number: row number str/int(1, 2, 3...)
            :param data: data to be written into the cell
        """
        try:
            self.work_book = openpyxl.load_workbook(excel_file)
            self.active_sheet = self.work_book[sheet] 
            # Get the column index based on the column name
            column_index = column_index_from_string(column_name)

            # Write the cell value
            self.active_sheet.cell(row=int(row_number), column=column_index, value=data)
            self.work_book.save(excel_file)
            get_cell_data = self.get_cell_value(excel_file, column_name, row_number, sheet)
            self.work_book.close()
            # verifying whether data is written or not
            if get_cell_data == data:
                self.logger.log_info("Cell value written successfully\n")
                return True
            else:
                self.logger.log_error("Unable to write into the cell\n")
                return False
        except Exception as error:
            self.logger.log_error(f"{error}\n")
            return False

    def write_file(self, excel_file, sheet, headers, data):
        """
            This method helps in writing the file with provided data
            :param excel_file: Excel file path
            :param sheet: specify the sheet name
            :param headers: specify the headers(Ex: Name, Age...)
            :param data: data to be written(nested list)
        """
        try:
            self.work_book = openpyxl.load_workbook(excel_file)
            self.active_sheet = self.work_book[sheet]
            self.active_sheet.delete_rows(1, self.active_sheet.max_row)
            self.active_sheet.append(headers)
            for row in data:
                self.active_sheet.append(row)
            self.logger.log_info("File written successfully\n")
            self.work_book.save(excel_file)
            return True
        except Exception as error:
            self.logger.log_error(f"{error}\n")
            return False
    
    def update_excel(self, excel_file, sheet, cell_values):
        """
            This method helps in updating an existing Excel file values with new cell values.
            :param excel_file: Excel file
            :param sheet_name: specify the sheet name
            :param cell_values: Dictionary of new cell values. Ex: {cell_address('A1', 'B1'): new_value}.
            :return: returns True if executed else False
        """
        try:
            self.work_book = openpyxl.load_workbook(excel_file)
            self.active_sheet = self.work_book[sheet]

            # Update the cell values
            for cell_address, new_value in cell_values.items():
                self.active_sheet[cell_address] = new_value
            self.work_book.save(excel_file)

            # Verifying the update
            for key, value in cell_values.items():
                get_cell_data = self.get_cell_value(excel_file, key[0], key[-1], sheet)
                if get_cell_data == value:
                    self.logger.log_info("Excel file updated successfully.\n")
                    return True
                else:
                    self.logger.log_error("Failed to update Excel file.\n")
                    return False
        except Exception as e:
            self.logger.log_error(f"Error updating Excel file: {e}\n")
            return False

    def create_excel_file(self, file_path, sheet_name, headers=None, data=None):
        """
            This helps in creating a new Excel file and data can be added.
            :param file_path: Path for the new Excel file.
            :param sheet_name: Name of the sheet to be created.
            :param headers: List of header values for the sheet.
            :param data: Nested List of data rows to be written to the sheet.
        """
        try:
            self.work_book = openpyxl.Workbook()
            self.active_sheet = self.work_book.active
            self.active_sheet.title = sheet_name
            self.work_book.save(file_path)
            if headers and data:
                # Write headers to the first row
                self.active_sheet.append(headers)
                # Write data to the subsequent rows
                for row in data:
                    self.active_sheet.append(row)
                self.work_book.save(file_path)
            if os.path.exists(file_path):
                self.logger.log_info(f"Excel file '{file_path}' created successfully.\n")
                return True
            else:
                self.logger.log_error(f"Failed to create Excel file at '{file_path}'.\n")
                return False
        except Exception as error:
            self.logger.log_error(f"Error creating Excel file: {error}\n")
            return False

    def create_sheet(self, excel_file, sheet_name):
        """
            This method helps in creating new sheet in the specified Excel file.
            :param excel_file: Excel file path
            :param sheet: specify the sheet name
        """
        try:
            self.work_book = openpyxl.load_workbook(excel_file)
            # Check if the sheet already exists
            if sheet_name in self.work_book.sheetnames:
                self.logger.log_error(f"Sheet '{sheet_name}' already exists in the workbook.\n")
                return False
            else:
                # Create a new sheet
                self.work_book.create_sheet(title=sheet_name)
                self.logger.log_info(f"New sheet '{sheet_name}' created successfully.\n")
                self.work_book.save(excel_file)
                if sheet_name in self.work_book.sheetnames:
                    return True
                else:
                    return False 
        except Exception as error:
            self.logger.log_error(f"Error: {error}\n")

    def get_sheets(self, excel_file):
        """
            This helps in getting all the sheets present in the specified excel file
            :param excel_file: Excel file path
            :return: returns list of sheet names
        """
        try:
            self.work_book = openpyxl.load_workbook(excel_file)
            return self.work_book.sheetnames
        except Exception as e:
            self.logger.log_error(f"Error: {e}\n")
            return False
    
    def delete_cell(self, excel_file, sheet, column_name, row_number):
        """
            This method helps in deleting the cell value from mentioned cell(row number & column name).
            :param excel_file: Excel file path
            :param sheet: specify the sheet name
            :param column_name: column name (A, B, C...)
            :param row_number: row number str/int(1, 2, 3...)
        """
        try:
            self.work_book = openpyxl.load_workbook(excel_file)
            self.active_sheet = self.work_book[sheet]
            column_index = column_index_from_string(column_name)

            # Set the cell's value to None to remove its content
            self.active_sheet.cell(row=row_number, column=column_index).value = None
            self.work_book.save(excel_file)

            # verifying whether data is written or not
            get_cell_data = self.get_cell_value(excel_file, column_name, row_number, sheet)
            if get_cell_data == None:
                self.logger.log_info(f'Cell at column "{column_name}", row "{row_number}" removed successfully.\n')
                return True
            else:
                self.logger.log_error("Unable to delete the cell\n")
                return False
        except Exception as error:
            self.logger.log_error(f"Error: {error}\n")
            return False

    def delete_row(self, excel_file, sheet, row_number):
        """
            This method helps in deleting an entire mentioned row.
            :param excel_file: Excel file path
            :param sheet: specify the sheet name
            :param row_number: row number str/int(1, 2, 3...)
        """
        try:
            self.work_book = openpyxl.load_workbook(excel_file)
            self.active_sheet = self.work_book[sheet]

            # Delete the entire row at the specified index
            self.active_sheet.delete_rows(row_number)
            self.work_book.save(excel_file)
            self.logger.log_info(f"Row {row_number} deleted successfully.\n")
            return True
        except Exception as error:
            self.logger.log_error(f"Error: {error}\n")
            return False

    def delete_column(self, excel_file, sheet, column_name):
        """
            This method helps in deleting an entire mentioned column.
            :param excel_file: Excel file path
            :param sheet: specify the sheet name
            :param column_name: column name (A, B, C...)
        """
        try:
            self.work_book = openpyxl.load_workbook(excel_file)
            self.active_sheet = self.work_book[sheet]
            column_index = column_index_from_string(column_name)
            # Delete the entire column at the specified index
            self.active_sheet.delete_cols(column_index)
            self.work_book.save(excel_file)
            self.logger.log_info(f"Column {column_name} deleted successfully.\n")
            return True
        except Exception as error:
            self.logger.log_error(f"Error: {error}\n")
            return False

    def delete_sheet(self, excel_file, sheet):
        """
            This method helps in deleting metioned/specified sheet in the Excel file.
            :param excel_file: Excel file path
            :param sheet: specify the sheet name
        """
        try:
            self.work_book = openpyxl.load_workbook(excel_file)
            # Check if the sheet exists before attempting to delete
            if sheet in self.work_book.sheetnames:
                self.active_sheet = self.work_book[sheet]
                self.work_book.remove(self.active_sheet)
                self.work_book.save(excel_file)
                if sheet not in self.work_book.sheetnames:
                    self.logger.log_info(f"Sheet '{sheet}' deleted successfully.\n")
                    return True
                else:
                    self.logger.log_error("Unable to delete the Sheet.\n")
                    return False
            else:
                self.logger.log_error(f"Sheet '{sheet}' does not exist.\n")
                return False
        except Exception as error:
            self.logger.log_error(f"Error: {error}\n")
            return False

    def excel_to_xml(self, excel_file, sheet=None, xml_file=None):
        """
            This method helps in converting excel into xml file/data
            :param excel_file: Excel file path
            :param sheet: specify the sheet name else active sheet will be picked
            :param xml_file: XML file path/name
            :return: creates xml file if xml file name/path is given else xml data is returned
        """
        try:
            # Read Excel file into a pandas DataFrame
            df = pd.read_excel(excel_file, sheet_name=sheet) if sheet else pd.read_excel(excel_file)
            # create root element of the XML tree
            root = ET.Element("data")

            # Iterate over each row in the DataFrame
            for _, row in df.iterrows():
                # Create a new XML element for each row
                item = ET.SubElement(root, "item")
                # Iterating over each column in the row
                for col_name, col_value in row.items():
                    # creating a new XML subelement for each column
                    col_elem = ET.SubElement(item, col_name)
                    col_elem.text = str(col_value)
            # updating in xml format
            rough_string = ET.tostring(root, 'utf-8')
            reparsed = minidom.parseString(rough_string)
            xml_string= reparsed.toprettyxml(indent="  ")

            # writing XML to file
            if xml_file:
                open(xml_file, "w", encoding="utf-8").write(xml_string)
                if os.path.exists(xml_file):
                    self.logger.log_info("XML File created at path specified.\n")
                    return True
                else:
                    return False
            else:
                self.logger.log_info("Converted to XML\n")
                return xml_string
        except Exception as error:
            self.logger.log_error(f"{error}\n")
            return False

    def excel_to_json(self, excel_file, sheet=None, json_file=None):
        """
            This method helps in converting excel into json file/data
            :param excel_file: Excel file path
            :param sheet: specify the sheet name else active sheet will be picked
            :param json_file: JSON file path/name
            :return: creates json file if json file path/name is given else json data is returned
        """
        try:
            df = pd.read_excel(excel_file, sheet_name=sheet) if sheet else pd.read_excel(excel_file)
            # Convert the DataFrame to a dictionary (records format)
            data = df.to_dict(orient='records')
            df = df.apply(lambda x: x.astype(str) if x.dtypes == 'object' or pd.api.types.is_datetime64_any_dtype(x) else x)
            data = df.to_dict(orient='records')

            # Write the dictionary to the JSON file
            if json_file:
                with open(json_file, 'w') as json_file:
                    json.dump(data, json_file, indent=4)
                    self.logger.log_info("JSON File created\n")
                return True
            else:
                self.logger.log_info("Converted to JSON\n")
                return data
        except Exception as error:
            self.logger.log_error(f"Error: {error}\n")
            return False
